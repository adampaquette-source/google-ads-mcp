"""Build the weekly ToolUp Top-200-SKUs ad-spend efficiency workbook.

Joins Google Ads product performance (last 30 days) with Shopify net sales
(last 30 days) for the ToolUp store, keeps the top 200 SKUs by ad spend, adds
efficiency columns (CVR, ROAS, Sales/Spend), sorts most-inefficient-first
(lowest ROAS), and writes a dated .xlsx into the project folder.

This script does NOT call any API. The scheduled routine pulls the two data
sources via MCP first, then passes their saved file paths in:

  uv run --with openpyxl python scripts/build_toolup_efficiency_report.py \
      --ads-json   /path/to/get_google_ads_product_performance-*.txt \
      --shopify-csv /path/to/query_sales_*_toolupstore_product_title_30d.csv

Optional:
  --out-dir   defaults to the project root (parent of this scripts/ dir)
  --date      YYYY-MM-DD label for the filename/header (defaults to today)

Matching is by product title (exact, then normalized on alphanumerics). SKUs
with no Shopify match had no sales in the window: they are written as $0 net
sales, shaded red, and annotated with a cell comment.
"""
import argparse
import csv
import json
import math
import re
from datetime import date as _date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

CUSTOMER_ID = "1864748540"  # ToolUp
TOP_N = 200
FONT = "Arial"


def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", s.lower())


def aggregate_ads(ads_json_path: str):
    """Sum cost/clicks/conversions/value per product_id, return top N by cost."""
    with open(ads_json_path) as f:
        data = json.load(f)
    rows = data["result"] if isinstance(data, dict) else data
    agg = {}
    for r in rows:
        pid = r["product_id"]
        a = agg.setdefault(pid, {"title": "", "cost": 0.0, "clicks": 0, "conv": 0.0, "cv": 0.0})
        if r.get("title"):
            a["title"] = r["title"]
        a["cost"] += r.get("cost", 0) or 0
        a["clicks"] += r.get("clicks", 0) or 0
        a["conv"] += r.get("conversions", 0) or 0
        a["cv"] += r.get("conversions_value", 0) or 0
    return sorted(agg.items(), key=lambda kv: -kv[1]["cost"])[:TOP_N]


def load_shopify(shopify_csv_path: str):
    """Map product title (raw and normalized) to (net_sales, orders)."""
    exact, norm = {}, {}
    with open(shopify_csv_path, newline="") as f:
        for row in csv.DictReader(f):
            title = row["Product title"].strip()
            val = (float(row["Net sales"] or 0), int(row["Orders"] or 0))
            exact[title] = val
            norm[_norm(title)] = val
    return exact, norm


def build(ads_json_path, shopify_csv_path, out_dir, label_date):
    top = aggregate_ads(ads_json_path)
    exact, norm = load_shopify(shopify_csv_path)

    rows = []
    for pid, a in top:
        title = a["title"].strip()
        if title in exact:
            ns, ords, nosale = exact[title][0], exact[title][1], False
        elif _norm(title) in norm:
            ns, ords, nosale = norm[_norm(title)][0], norm[_norm(title)][1], False
        else:
            ns, ords, nosale = 0.0, 0, True
        cost, clicks, conv, cv = a["cost"], a["clicks"], a["conv"], a["cv"]
        cvr = (conv / clicks) if clicks else 0.0
        roas = (cv / cost) if cost else 0.0
        sps = (ns / cost) if cost else 0.0
        rows.append((pid, title, cost, clicks, conv, cvr, cv, roas, ns, ords, sps, nosale))
    rows.sort(key=lambda x: (x[7], -x[2]))  # ROAS asc, spend desc

    wb = Workbook()
    ws = wb.active
    ws.title = "Top 200 by Spend"
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    base_font = Font(name=FONT, size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    flag_fill = PatternFill("solid", fgColor="FCE4E4")

    ws["A1"] = "ToolUp - Top 200 SKUs by Google Ads Spend (Last 30 Days)"
    ws["A1"].font = Font(name=FONT, bold=True, size=13)
    ws["A2"] = (
        "Sorted by ad-spend inefficiency: lowest ROAS first. ROAS = ad-attributed conv value / ad spend. "
        "Sales/Spend = total Shopify net sales / ad spend (all channels). Rows shaded red = $0 Shopify "
        f"sales in last 30 days. Account {CUSTOMER_ID}. Values as of {label_date}."
    )
    ws["A2"].font = Font(name=FONT, italic=True, size=9, color="595959")

    headers = ["Rank", "SKU", "Product Title", "Ad Spend", "Clicks", "Ad Conv", "CVR",
               "Ad Conv Value", "ROAS", "Shopify Net Sales (30d)", "Shopify Orders (30d)", "Sales / Spend"]
    HROW = 4
    for c, h in enumerate(headers, 1):
        cell = ws.cell(HROW, c, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    money2 = "$#,##0.00"
    money0 = '$#,##0;($#,##0);"-"'
    pct = "0.0%"
    mult = "0.0x"
    num1 = "0.0"
    start = HROW + 1
    for i, (pid, title, cost, clicks, conv, cvr, cv, roas, ns, ords, sps, nosale) in enumerate(rows):
        rr = start + i
        ws.cell(rr, 1, i + 1)
        ws.cell(rr, 2, pid)
        ws.cell(rr, 3, title)
        ws.cell(rr, 4, round(cost, 2)).number_format = money2
        ws.cell(rr, 5, clicks)
        ws.cell(rr, 6, round(conv, 1)).number_format = num1
        ws.cell(rr, 7, round(cvr, 4)).number_format = pct
        ws.cell(rr, 8, round(cv, 2)).number_format = money0
        ws.cell(rr, 9, round(roas, 2)).number_format = mult
        jc = ws.cell(rr, 10, round(ns, 2))
        jc.number_format = money0
        ws.cell(rr, 11, ords)
        ws.cell(rr, 12, round(sps, 2)).number_format = mult
        if nosale:
            jc.comment = Comment("No Shopify sales in last 30 days (not among store sellers in period).", "Ads MCP")
        for c in range(1, 13):
            cell = ws.cell(rr, c)
            cell.font = base_font
            cell.border = border
            if nosale:
                cell.fill = flag_fill

    last = start + len(rows) - 1
    tr = last + 1
    tot_cost = sum(x[2] for x in rows)
    tot_clicks = sum(x[3] for x in rows)
    tot_conv = sum(x[4] for x in rows)
    tot_cv = sum(x[6] for x in rows)
    tot_ns = sum(x[8] for x in rows)
    tot_ord = sum(x[9] for x in rows)
    totals = [
        ("", None), ("", None), ("TOTAL / PORTFOLIO (top 200)", None),
        (round(tot_cost, 2), money2), (tot_clicks, None), (round(tot_conv, 1), num1),
        (round(tot_conv / tot_clicks, 4) if tot_clicks else 0, pct),
        (round(tot_cv, 2), money0),
        (round(tot_cv / tot_cost, 2) if tot_cost else 0, mult),
        (round(tot_ns, 2), money0), (tot_ord, None),
        (round(tot_ns / tot_cost, 2) if tot_cost else 0, mult),
    ]
    for c, (v, fmt) in enumerate(totals, 1):
        cell = ws.cell(tr, c, v)
        if fmt:
            cell.number_format = fmt
        cell.font = Font(name=FONT, bold=True, size=10)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")
        cell.border = border

    for i, w in enumerate([6, 11, 52, 12, 9, 9, 8, 15, 8, 18, 16, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{start}"
    ws.auto_filter.ref = f"A{HROW}:L{last}"
    for r in range(start, last + 1):
        ws.row_dimensions[r].height = 28

    out_path = Path(out_dir) / f"ToolUp_Top200_SKUs_AdSpend_Efficiency_{label_date}.xlsx"
    wb.save(out_path)

    bad = sum(
        1 for x in rows
        if any(isinstance(v, float) and (math.isnan(v) or math.isinf(v)) for v in (x[5], x[7], x[10]))
    )
    nosale_n = sum(1 for x in rows if x[11])
    return {
        "out_path": str(out_path),
        "rows": len(rows),
        "nosale_flagged": nosale_n,
        "bad_numeric_cells": bad,
        "tot_cost": tot_cost,
        "tot_cv": tot_cv,
        "tot_ns": tot_ns,
        "roas": (tot_cv / tot_cost) if tot_cost else 0,
        "sales_per_spend": (tot_ns / tot_cost) if tot_cost else 0,
    }


def main():
    project_root = Path(__file__).resolve().parent.parent
    p = argparse.ArgumentParser()
    p.add_argument("--ads-json", required=True, help="Saved get_google_ads_product_performance result file")
    p.add_argument("--shopify-csv", required=True, help="Saved shopify_query_sales CSV (product_title, 30d)")
    p.add_argument("--out-dir", default=str(project_root), help="Output directory (default: project root)")
    p.add_argument("--date", default=_date.today().isoformat(), help="YYYY-MM-DD label (default: today)")
    args = p.parse_args()

    res = build(args.ads_json, args.shopify_csv, args.out_dir, args.date)
    if res["bad_numeric_cells"]:
        raise SystemExit(f"ERROR: {res['bad_numeric_cells']} bad numeric cells; aborting.")
    print("WROTE", res["out_path"])
    print(f"rows={res['rows']} no-sale flagged={res['nosale_flagged']} bad_cells={res['bad_numeric_cells']}")
    print(
        f"portfolio: spend ${res['tot_cost']:,.0f} | ad conv val ${res['tot_cv']:,.0f} | "
        f"ROAS {res['roas']:.2f}x | shopify sales ${res['tot_ns']:,.0f} | sales/spend {res['sales_per_spend']:.2f}x"
    )


if __name__ == "__main__":
    main()
